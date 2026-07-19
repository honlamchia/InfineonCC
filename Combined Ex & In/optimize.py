"""
optimize.py
===========
Internal route optimiser (PuLP) with the brief's WeightedScore in the objective.

Two objective modes (cost/kg must enter the score BEFORE the solve, not after):

  official (cost/kg)  - the graded MinScore at DELIVERY grain: cost/kg = BaseCostEUR /
                        individual delivery ChargeableWeight_KG (benchmark definition,
                        225/225). 225 delivery scores share 132 internal route decisions.
                        Score = 0.4*norm(BaseLeadTimeDays) + 0.4*norm(CostPerKG)
                                + 0.2*norm(RiskScore).

  proxy (cost/piece)  - Internal Routing Proxy Score for ALL 240 shipments, used for
                        the resilience / scenario-swap story. cost = BaseCostEUR / Qty
                        (cost per piece, an honest proxy where no weight exists).
                        Same 40/40/20 shape, labelled a PROXY - not the official MinScore.

In BOTH modes the LEAD term uses BaseLeadTimeDays (the official definition).
EffectiveLeadTimeDays and capacity remain CONSTRAINTS / reporting only - multi-week
throughput and the 12-week horizon gate feasibility but do not enter the score, unless
the judges approve substituting effective lead.
"""

from pathlib import Path
import argparse
import math
import pandas as pd
import pulp

import build_candidates as bc

W_LEAD, W_COST, W_RISK = 0.40, 0.40, 0.20
UNASSIGNED_PENALTY = 10.0
DAYS_PER_WEEK = 7
SCENARIOS = ["Normal", "PrimaryHubDown", "AirCapacityReduced"]
EXPEDITE_LEAD_DAYS = 6
PLANNING_HORIZON_WEEKS = 12


# --------------------------------------------------------------------------
# Weights / cost basis
# --------------------------------------------------------------------------
def add_capacity_fields(cand: pd.DataFrame, qty: dict, horizon_weeks: int = None) -> pd.DataFrame:
    if horizon_weeks is None:
        horizon_weeks = PLANNING_HORIZON_WEEKS
    c = cand.copy()
    c["Qty"] = c["ShipmentID"].map(qty)
    c["BottleneckCapacityPerWeek"] = c[
        ["CapacityUnitsPerWeek", "orig_remaining_capacity_units",
         "dest_remaining_capacity_units"]].min(axis=1)
    c["WeeksRequired"] = [
        (10**9 if b <= 0 else max(1, math.ceil(q / b)))
        for q, b in zip(c["Qty"], c["BottleneckCapacityPerWeek"])]
    c["MultiWeek"] = (c["WeeksRequired"] > 1) & (c["BottleneckCapacityPerWeek"] > 0)
    c["EffectiveLeadTimeDays"] = c["BaseLeadTimeDays"] + DAYS_PER_WEEK * (
        c["WeeksRequired"].clip(upper=520) - 1)
    c["WeeklyFootprint"] = c[["Qty", "BottleneckCapacityPerWeek"]].min(axis=1).clip(lower=0)

    # Horizon = throughput weeks to clear capacity (transit time NOT counted).
    c["Escalation"] = (c["BottleneckCapacityPerWeek"] > 0) & (c["WeeksRequired"] > horizon_weeks)
    c["CapacityReason"] = "ok"
    c.loc[c["BottleneckCapacityPerWeek"] <= 0, "CapacityReason"] = "zero-capacity hub/route"
    c.loc[c["Escalation"], "CapacityReason"] = f"capacity escalation (> {horizon_weeks} throughput wks)"
    return c


def feasible_for_model(c: pd.DataFrame) -> pd.DataFrame:
    return c[(c["BottleneckCapacityPerWeek"] > 0) & (~c["Escalation"])].copy()


def add_cost_basis(c: pd.DataFrame) -> pd.DataFrame:
    """Proxy cost term: cost per piece (the official cost/kg lives in the
    delivery-grain functions below - the old aggregate-weight kg mode is removed)."""
    c = c.copy()
    c["CostForScore"] = c["BaseCostEUR"] / c["Qty"]
    return c


def objective_universe(sheets: dict) -> list:
    return sheets["internal"]["ShipmentID"].tolist()                           # all 240


# --------------------------------------------------------------------------
# Scoring
# --------------------------------------------------------------------------
def fit_scaler(sheets, horizon_weeks) -> dict:
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    frames = []
    for sc in SCENARIOS:
        cand = add_capacity_fields(bc.apply_capability(bc.build_candidates(sheets, sc)),
                                   qty, horizon_weeks)
        cand = add_cost_basis(cand)
        frames.append(feasible_for_model(cand))
    u = pd.concat(frames, ignore_index=True)
    return {"lead": (u["BaseLeadTimeDays"].min(), u["BaseLeadTimeDays"].max()),
            "cost": (u["CostForScore"].min(), u["CostForScore"].max()),
            "risk": (u["RiskScore"].min(), u["RiskScore"].max())}


def _scale(s, lo, hi):
    return 0.0 if hi == lo else (s - lo) / (hi - lo)


def add_weighted_score(c: pd.DataFrame, scaler: dict) -> pd.DataFrame:
    c = c.copy()
    c["n_lead"] = _scale(c["BaseLeadTimeDays"], *scaler["lead"])   # official lead term
    c["n_cost"] = _scale(c["CostForScore"],     *scaler["cost"])
    c["n_risk"] = _scale(c["RiskScore"],        *scaler["risk"])
    c["WeightedScore"] = W_LEAD * c["n_lead"] + W_COST * c["n_cost"] + W_RISK * c["n_risk"]
    return c


def iso_week(ship_dates: pd.Series) -> pd.Series:
    d = pd.to_datetime(ship_dates, errors="coerce")
    iso = d.dt.isocalendar()
    return iso["year"].astype(str) + "-W" + iso["week"].astype(str).str.zfill(2)


# --------------------------------------------------------------------------
# Model
# --------------------------------------------------------------------------
def solve(cand: pd.DataFrame, sheets: dict, scaler: dict, universe: list) -> dict:
    cand = add_weighted_score(cand, scaler).reset_index(drop=True)
    cand["week"] = iso_week(cand["ShipDate"])
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    hub_remaining = sheets["hub"].set_index("HubID")["remaining_capacity_units"].to_dict()

    # only candidates for shipments in this objective universe
    cand = cand[cand["ShipmentID"].isin(universe)].reset_index(drop=True)
    rows = cand.index.tolist()

    model = pulp.LpProblem("route_optimiser", pulp.LpMinimize)
    x = {i: pulp.LpVariable(f"x_{i}", cat="Binary") for i in rows}
    u = {s: pulp.LpVariable(f"u_{s}", cat="Binary") for s in universe}

    model += (pulp.lpSum(cand.loc[i, "WeightedScore"] * x[i] for i in rows)
              + UNASSIGNED_PENALTY * pulp.lpSum(u[s] for s in universe))

    rows_by_ship = cand.groupby("ShipmentID").groups
    for s in universe:
        model += (pulp.lpSum(x[i] for i in rows_by_ship.get(s, [])) + u[s] == 1, f"assign_{s}")

    for (rid, wk), grp in cand.groupby(["RouteOptionID", "week"]):
        cap = grp["CapacityUnitsPerWeek"].iloc[0]
        model += (pulp.lpSum(cand.loc[i, "WeeklyFootprint"] * x[i] for i in grp.index) <= cap,
                  f"routecap_{rid}_{wk}")

    usage: dict = {}
    for i in rows:
        r = cand.loc[i]
        for hub in {r["FromHub"], r["ToHub"]}:
            usage.setdefault((hub, r["week"]), []).append((r["WeeklyFootprint"], x[i]))
    for (hub, wk), terms in usage.items():
        cap = hub_remaining.get(hub, float("inf"))
        model += (pulp.lpSum(w * v for w, v in terms) <= cap, f"hubcap_{hub}_{wk}")

    model.solve(pulp.PULP_CBC_CMD(msg=False))
    chosen = cand.loc[[i for i in rows if x[i].value() == 1]].copy()
    chosen["Assigned"] = True
    unassigned_ids = [s for s in universe if u[s].value() == 1]
    return {"status": pulp.LpStatus[model.status], "chosen": chosen,
            "unassigned_ids": unassigned_ids, "candidates": cand}


def primary_lane_baseline(cand, sheets, scaler, universe) -> dict:
    prim = cand[cand["IsPrimary"] == "Yes"].copy()
    if not len(prim):
        return {"by_ship": pd.Series(dtype="float64"), "served": set()}
    res = solve(prim, sheets, scaler, universe)
    by_ship = res["chosen"].set_index("ShipmentID")["WeightedScore"]
    return {"by_ship": by_ship, "served": set(by_ship.index)}


def _r(v):
    return None if v is None or pd.isna(v) else round(float(v), 4)


def penalised_objective(score_by_ship, n_total):
    return round(score_by_ship.sum() + UNASSIGNED_PENALTY * (n_total - score_by_ship.notna().sum()), 4)


def run_scenario(scenario, sheets, scaler, horizon_weeks) -> dict:
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    universe = objective_universe(sheets)
    n = len(universe)

    annotated = add_cost_basis(
        add_capacity_fields(bc.apply_capability(bc.build_candidates(sheets, scenario)),
                            qty, horizon_weeks))
    annotated = annotated[annotated["ShipmentID"].isin(universe)]
    cand = feasible_for_model(annotated)
    res = solve(cand, sheets, scaler, universe)

    opt_by_ship = res["chosen"].set_index("ShipmentID")["WeightedScore"]
    base = primary_lane_baseline(cand, sheets, scaler, universe)

    cap_feasible = set(annotated["ShipmentID"])
    within_cap = set(annotated.loc[annotated["BottleneckCapacityPerWeek"] > 0, "ShipmentID"])
    within_horiz = set(cand["ShipmentID"])
    real = annotated[annotated["BottleneckCapacityPerWeek"] > 0]
    best_real = real.sort_values("EffectiveLeadTimeDays").groupby("ShipmentID").first()
    ACTIONS = {
        "handling_infeasible": "No handling-compatible hub - source alternate hub / supplier exception",
        "zero_capacity":       "Only routes pass a zero-capacity hub - free capacity or reroute",
        "capacity_escalation": "Exceeds quarter horizon - split order across periods or add capacity",
        "capacity_contention": "Weekly capacity contention - reschedule to another week"}
    reasons = {k: 0 for k in ACTIONS}
    reason_by_ship, unassigned_detail = {}, []
    for sid in res["unassigned_ids"]:
        if sid not in cap_feasible:   rk = "handling_infeasible"
        elif sid not in within_cap:   rk = "zero_capacity"
        elif sid not in within_horiz: rk = "capacity_escalation"
        else:                         rk = "capacity_contention"
        reasons[rk] += 1
        reason_by_ship[sid] = rk
        bw = int(best_real.loc[sid, "WeeksRequired"]) if sid in best_real.index else None
        br = best_real.loc[sid, "RouteOptionID"] if sid in best_real.index else None
        unassigned_detail.append({"scenario": scenario, "ShipmentID": sid, "Reason": rk,
                                  "BestAvailableWeeks": bw, "BestAvailableRoute": br,
                                  "RecommendedAction": ACTIONS[rk]})

    common = set(opt_by_ship.index) & base["served"]
    summary = {
        "scenario": scenario, "objective": "proxy cost/piece", "universe": n, "solver": res["status"],
        "opt_solved": int(opt_by_ship.notna().sum()), "opt_unassigned": len(res["unassigned_ids"]),
        "opt_avg_MinScore": _r(opt_by_ship.mean()), "opt_median": _r(opt_by_ship.median()),
        "opt_std": _r(opt_by_ship.std()), "opt_min": _r(opt_by_ship.min()), "opt_max": _r(opt_by_ship.max()),
        "base_solved": len(base["served"]),
        "base_avg": _r(base["by_ship"].mean()) if len(base["served"]) else None,
        "common_n": len(common),
        "opt_avg_common": _r(opt_by_ship.reindex(common).mean()) if common else None,
        "base_avg_common": _r(base["by_ship"].reindex(common).mean()) if common else None,
        "opt_penalised_obj": penalised_objective(opt_by_ship, n),
        "base_penalised_obj": penalised_objective(base["by_ship"], n),
        "multiweek_selected": int(res["chosen"]["MultiWeek"].sum()),
        "horizon_weeks": horizon_weeks,
        **{f"unassigned_{k}": v for k, v in reasons.items()}}
    return {"summary": summary, "result": res,
            "reason_by_ship": reason_by_ship, "unassigned_detail": unassigned_detail}



# ==========================================================================
# OFFICIAL score at DELIVERY grain (cost/kg = BaseCostEUR / per-delivery weight)
# One route decision per internal shipment; 225 delivery-level scores.
# ==========================================================================
def deliveries_by_shipment(sheets):
    E = pd.read_excel(bc.WORKBOOK, sheet_name="External Shipments")
    dbs = {}
    for _, r in E.iterrows():
        dbs.setdefault(r["InternalShipmentID_Link"], []).append(
            (r["DeliveryNo"], float(r["ChargeableWeight_KG"])))
    return dbs, E


def expand_deliveries(cand, dbs):
    """One row per (candidate route, linked delivery), with per-delivery cost/kg."""
    rows = []
    for idx, r in cand.iterrows():
        for dno, w in dbs.get(r["ShipmentID"], []):
            rows.append({"cand_idx": idx, "ShipmentID": r["ShipmentID"],
                         "RouteOptionID": r["RouteOptionID"], "DeliveryNo": dno,
                         "ChargeableWeight_KG": w,
                         "BaseLeadTimeDays": r["BaseLeadTimeDays"],
                         "BaseCostEUR": r["BaseCostEUR"],
                         "RiskScore": r["RiskScore"],
                         "CostPerKG": r["BaseCostEUR"] / w if w > 0 else float("nan")})
    return pd.DataFrame(rows)


def fit_scaler_delivery(sheets, horizon_weeks, dbs):
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    linked = set(dbs)
    frames = []
    for sc in SCENARIOS:
        cand = feasible_for_model(add_capacity_fields(
            bc.apply_capability(bc.build_candidates(sheets, sc)), qty, horizon_weeks))
        cand = cand[cand["ShipmentID"].isin(linked)]
        frames.append(expand_deliveries(cand, dbs))
    u = pd.concat(frames, ignore_index=True)
    return {"lead": (u["BaseLeadTimeDays"].min(), u["BaseLeadTimeDays"].max()),
            "cost": (u["CostPerKG"].min(), u["CostPerKG"].max()),
            "risk": (u["RiskScore"].min(), u["RiskScore"].max())}


def _score_expansion(ex, scaler):
    if ex.empty:
        return pd.DataFrame(columns=["cand_idx", "ShipmentID", "RouteOptionID", "DeliveryNo",
                                     "ChargeableWeight_KG", "BaseLeadTimeDays", "BaseCostEUR",
                                     "RiskScore", "CostPerKG", "n_lead", "n_cost", "n_risk",
                                     "DeliveryMinScore"])
    ex = ex.copy()
    ex["n_lead"] = _scale(ex["BaseLeadTimeDays"], *scaler["lead"])
    ex["n_cost"] = _scale(ex["CostPerKG"],        *scaler["cost"])
    ex["n_risk"] = _scale(ex["RiskScore"],        *scaler["risk"])
    ex["DeliveryMinScore"] = (W_LEAD * ex["n_lead"] + W_COST * ex["n_cost"]
                              + W_RISK * ex["n_risk"])
    return ex


def solve_official_delivery(scenario, sheets, scaler, horizon_weeks, dbs,
                           restrict_primary=False):
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    hub_remaining = sheets["hub"].set_index("HubID")["remaining_capacity_units"].to_dict()
    linked = set(dbs)

    cand = feasible_for_model(add_capacity_fields(
        bc.apply_capability(bc.build_candidates(sheets, scenario)), qty, horizon_weeks))
    cand = cand[cand["ShipmentID"].isin(linked)].copy()
    if restrict_primary:                       # baseline: primary planned lanes only
        cand = cand[cand["IsPrimary"] == "Yes"].copy()
    cand["week"] = iso_week(cand["ShipDate"])

    # delivery-level scores -> per-candidate objective coefficient = SUM over its deliveries
    ex = _score_expansion(expand_deliveries(cand, dbs), scaler)
    coeff = ex.groupby("cand_idx")["DeliveryMinScore"].sum().to_dict()

    shipments = [s for s in sheets["internal"]["ShipmentID"] if s in linked]
    n_deliv = {s: len(dbs[s]) for s in shipments}
    rows = [i for i in cand.index if i in coeff]

    model = pulp.LpProblem("official_delivery", pulp.LpMinimize)
    x = {i: pulp.LpVariable(f"x_{i}", cat="Binary") for i in rows}
    u = {s: pulp.LpVariable(f"u_{s}", cat="Binary") for s in shipments}

    # minimise total delivery-level score; unassigned penalised per unserved delivery
    model += (pulp.lpSum(coeff[i] * x[i] for i in rows)
              + UNASSIGNED_PENALTY * pulp.lpSum(n_deliv[s] * u[s] for s in shipments))

    rows_by_ship = cand.loc[rows].groupby("ShipmentID").groups
    for s in shipments:
        model += (pulp.lpSum(x[i] for i in rows_by_ship.get(s, [])) + u[s] == 1, f"assign_{s}")

    # capacity charged ONCE per internal shipment (not per delivery)
    for (rid, wk), grp in cand.loc[rows].groupby(["RouteOptionID", "week"]):
        cap = grp["CapacityUnitsPerWeek"].iloc[0]
        model += (pulp.lpSum(cand.loc[i, "WeeklyFootprint"] * x[i] for i in grp.index) <= cap,
                  f"routecap_{rid}_{wk}")
    usage = {}
    for i in rows:
        r = cand.loc[i]
        for hub in {r["FromHub"], r["ToHub"]}:
            usage.setdefault((hub, r["week"]), []).append((r["WeeklyFootprint"], x[i]))
    for (hub, wk), terms in usage.items():
        cap = hub_remaining.get(hub, float("inf"))
        model += (pulp.lpSum(w * v for w, v in terms) <= cap, f"hubcap_{hub}_{wk}")

    model.solve(pulp.PULP_CBC_CMD(msg=False))

    chosen_idx = [i for i in rows if x[i].value() == 1]
    chosen = cand.loc[chosen_idx].copy()
    chosen_route = chosen.set_index("ShipmentID")["RouteOptionID"].to_dict()
    if len(ex):
        served_ex = ex[ex.apply(lambda r: chosen_route.get(r["ShipmentID"]) == r["RouteOptionID"], axis=1)]
    else:
        served_ex = ex
    unassigned_ids = [s for s in shipments if u[s].value() == 1]

    # why each linked shipment is unassigned (same taxonomy as the proxy run)
    annotated = add_capacity_fields(
        bc.apply_capability(bc.build_candidates(sheets, scenario)), qty, horizon_weeks)
    annotated = annotated[annotated["ShipmentID"].isin(linked)]
    cap_feasible = set(annotated["ShipmentID"])
    within_cap = set(annotated.loc[annotated["BottleneckCapacityPerWeek"] > 0, "ShipmentID"])
    within_horiz = set(cand["ShipmentID"])
    reason_by_ship = {}
    for sid in unassigned_ids:
        if sid not in cap_feasible:   reason_by_ship[sid] = "handling_infeasible"
        elif sid not in within_cap:   reason_by_ship[sid] = "zero_capacity"
        elif sid not in within_horiz: reason_by_ship[sid] = "capacity_escalation"
        else:                         reason_by_ship[sid] = "capacity_contention"
    return {"status": pulp.LpStatus[model.status], "chosen": chosen,
            "chosen_route": chosen_route, "delivery_scores": served_ex,
            "shipments": shipments, "unassigned_ids": unassigned_ids,
            "reason_by_ship": reason_by_ship}


ASSUMPTIONS_OFFICIAL = [
    ("STATUS: route extension, non-official for external",
     "Per Infineon instructor clarification, the OFFICIAL external MinScore is a direct 40/40/20 "
     "calculation on the supplied External Shipments.{BestLeadTimeDays, LowestCostPerKG_EUR, "
     "LowestRiskScore} columns (see optimizer_external_official_scores.xlsx). This combined "
     "delivery-grain model derives the external cost/kg from route_options and is therefore a "
     "ROUTE EXTENSION - its per-delivery score is a ScenarioRouteScore, not the official ExternalMinScore. "
     "The internal route optimiser remains valid; only the external scoring basis is superseded."),
    ("Objective (graded MinScore)",
     "Normalised 40/40/20 MinScore per DELIVERY: 0.4*norm(BaseLeadTimeDays) + 0.4*norm(CostPerKG) "
     "+ 0.2*norm(RiskScore). Lower is better. Min-max normalisation is a stated modelling assumption "
     "(the brief does not specify bounds); bounds shown below and used live by the formulas."),
    ("Cost/kg definition",
     "CostPerKG = route BaseCostEUR / individual delivery ChargeableWeight_KG - matches the dataset's "
     "own LowestCostPerKG_EUR benchmark definition (verified 225/225). 225 delivery scores share 132 "
     "upstream route decisions (one route per internal shipment; capacity charged once per shipment)."),
    ("Universe",
     "Official score covers the 132 weight-linked internal shipments / 225 deliveries. Internal legs "
     "carry no weight (UoM=ST), so cost/kg is undefined for the other 108 shipments - they are routed "
     "in the separate proxy (cost/piece) model. Averages are over scored deliveries; unassigned "
     "deliveries are listed in Unassigned, never silently dropped."),
    ("Scenario layers",
     "Two independent axes. (1) Route scenario (Normal/PrimaryHubDown/AirCapacityReduced) filters "
     "Route_Options only. (2) Hub disruption is a per-hub capacity cut in Hub_Constraints, activated "
     "via --hub-disruption and applied only to matching hubs (or every tagged hub in legacy 'all' "
     "mode). The two DisruptionScenario columns are never matched to each other. The exact axes used "
     "for THIS workbook are stated in the 'Active run configuration' rows below and in Summary."),
    ("Handling capability",
     "Cold chain + hazard class via dedicated hub booleans; BOTH origin and destination hubs must "
     "qualify (stated assumption)."),
    ("Capacity model",
     "*_pct fields are decimal fractions. remaining = weekly*(max_util*(1-reduction)) - current_load. "
     "bottleneck = min(route cap, origin remaining, dest remaining); bottleneck<=0 rejected. "
     "WeeklyFootprint = min(Qty, bottleneck) charged identically to route + both hubs per ship-week. "
     "Planning horizon: WeeksRequired > 12 throughput weeks -> capacity escalation, never a normal "
     "solution. Approximation: footprint charged in ship week (full weekly-flow f[s,r,w] is the "
     "documented next step)."),
    ("Baseline",
     "Primary planned lanes (IsPrimary=Yes) run through the SAME capacity-constrained solver, scored "
     "at the same delivery grain with the same normalisation. Q1/Q3 thresholds come from the baseline "
     "delivery-score distribution (per Hackathon_Guide). No primary lanes exist under the disruption "
     "scenarios - reported as such, not as a zero score."),
    ("Solver",
     "Binary PuLP/CBC model; one route or explicit unassigned slack per shipment (never an infeasible "
     "model)."),
]

TRADEOFF_OFFICIAL = (
    "Lead time and cost/kg carry equal 40% weights: the model only accepts a slower route when its "
    "cost/kg saving is at least proportional, and vice versa; risk (20%) breaks ties toward robust "
    "lanes. In practice light-weight deliveries have high cost/kg, so their shipments prefer cheap "
    "road/sea lanes unless lead time dominates; heavy deliveries dilute cost/kg and take faster "
    "lanes. Under PrimaryHubDown / AirCapacityReduced the candidate pool degrades, so the optimiser "
    "trades a small score increase for feasible lower-risk alternates instead of unavailable "
    "primaries. Handling rules remove non-compliant hubs before scoring; capacity and the 12-week "
    "horizon act as hard constraints, so no score wins by violating feasibility."
)


def _run_config_rows(hub_disruption, scenarios):
    """Explicit per-run context so a reader never has to infer the mode from the filename."""
    if hub_disruption is None:
        hub_txt = "None (clean network - no hub capacity cuts applied)"
    elif hub_disruption == bc.HUB_DISRUPTION_LEGACY_ALL:
        hub_txt = "all (LEGACY: every hub with a recorded reduction is cut, regardless of event)"
    else:
        hub_txt = (f"{hub_disruption} (CapacityReductionPct applied only to hubs tagged "
                   f"'{hub_disruption}')")
    return [
        ("Active route scenarios", ", ".join(scenarios) +
         "  (each scenario is a separate block in the score/Summary sheets)"),
        ("Active hub disruption", hub_txt +
         ".  This SAME hub mode is layered onto every route scenario in this workbook."),
    ]


def _export_official(out_path, scaler, per_scenario, unassigned_rows, route_frames,
                     hub_disruption=None, scenarios=None):
    """Submission-grade workbook: live formulas + assumptions + baseline + unassigned."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    AR = Font(name="Arial", size=10); BD = Font(name="Arial", size=10, bold=True)
    H1 = Font(name="Arial", size=14, bold=True)
    YEL = PatternFill("solid", fgColor="FFFF00"); GRY = PatternFill("solid", fgColor="D9D9D9")
    FMT = "0.0000"

    wb = Workbook()

    # ---------------- Assumptions (bounds live here) ----------------
    wa = wb.create_sheet("Assumptions")
    wa["A1"] = "ChainLab - Combined Route Extension (delivery-grain) - NON-OFFICIAL for external"; wa["A1"].font = H1
    r = 3
    # Active run configuration FIRST so the reader sees exactly what this file contains.
    wa.cell(r, 1, "Active run configuration").font = H1; r += 1
    for k, v in _run_config_rows(hub_disruption, scenarios or []):
        wa.cell(r, 1, k).font = BD
        c = wa.cell(r, 2, v); c.font = AR; c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    for k, v in ASSUMPTIONS_OFFICIAL:
        wa.cell(r, 1, k).font = BD
        c = wa.cell(r, 2, v); c.font = AR; c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    wa.cell(r, 1, "Normalisation bounds (min-max, fixed across scenarios)").font = BD; r += 1
    bounds_row = {}
    for name, key in [("Lead (days)", "lead"), ("Cost/kg (EUR)", "cost"), ("Risk", "risk")]:
        wa.cell(r, 1, name).font = AR
        lo, hi = wa.cell(r, 2, float(scaler[key][0])), wa.cell(r, 3, float(scaler[key][1]))
        for c in (lo, hi):
            c.font = AR; c.fill = YEL; c.number_format = FMT
        bounds_row[key] = r; r += 1
    B = {k: (f"Assumptions!$B${v}", f"Assumptions!$C${v}") for k, v in bounds_row.items()}
    r += 1
    wa.cell(r, 1, "Tradeoff explanation (lead time vs cost/kg vs risk)").font = BD; r += 1
    c = wa.cell(r, 1, TRADEOFF_OFFICIAL); c.font = AR
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wa.merge_cells(start_row=r, start_column=1, end_row=r + 7, end_column=6)
    wa.column_dimensions["A"].width = 24; wa.column_dimensions["B"].width = 100

    # ---------------- score sheets with live formulas ----------------
    COLS = ["scenario", "DeliveryNo", "ShipmentID", "RouteOptionID", "ChargeableWeight_KG",
            "BaseLeadTimeDays", "BaseCostEUR", "CostPerKG", "RiskScore",
            "n_lead", "n_cost", "n_risk", "DeliveryMinScore"]

    def write_scores(ws, frames):
        for j, h in enumerate(COLS, 1):
            c = ws.cell(1, j, h); c.font = BD; c.fill = GRY
        ws.freeze_panes = "A2"
        L = {h: get_column_letter(i + 1) for i, h in enumerate(COLS)}
        blocks, r = {}, 2
        for sc, df in frames:
            start = r
            for _, row in df.iterrows():
                for h in ("scenario", "DeliveryNo", "ShipmentID", "RouteOptionID",
                          "ChargeableWeight_KG", "BaseLeadTimeDays", "BaseCostEUR", "RiskScore"):
                    ws.cell(r, COLS.index(h) + 1,
                            sc if h == "scenario" else row[h]).font = AR
                ws.cell(r, COLS.index("CostPerKG") + 1,
                        f"={L['BaseCostEUR']}{r}/{L['ChargeableWeight_KG']}{r}")
                ws.cell(r, COLS.index("n_lead") + 1,
                        f"=({L['BaseLeadTimeDays']}{r}-{B['lead'][0]})/({B['lead'][1]}-{B['lead'][0]})")
                ws.cell(r, COLS.index("n_cost") + 1,
                        f"=({L['CostPerKG']}{r}-{B['cost'][0]})/({B['cost'][1]}-{B['cost'][0]})")
                ws.cell(r, COLS.index("n_risk") + 1,
                        f"=({L['RiskScore']}{r}-{B['risk'][0]})/({B['risk'][1]}-{B['risk'][0]})")
                ws.cell(r, COLS.index("DeliveryMinScore") + 1,
                        f"=0.4*{L['n_lead']}{r}+0.4*{L['n_cost']}{r}+0.2*{L['n_risk']}{r}")
                for h in ("CostPerKG", "n_lead", "n_cost", "n_risk", "DeliveryMinScore"):
                    cell = ws.cell(r, COLS.index(h) + 1); cell.font = AR; cell.number_format = FMT
                r += 1
            blocks[sc] = (start, r - 1)
        return blocks, L

    wd = wb.create_sheet("DeliveryScores")
    blocks_o, L = write_scores(wd, [(sc, p["ds"]) for sc, p in per_scenario.items()])
    wbs = wb.create_sheet("BaselineScores")
    base_frames = [(sc, p["base_ds"]) for sc, p in per_scenario.items() if len(p["base_ds"])]
    blocks_b, _ = write_scores(wbs, base_frames)
    wbs.cell(1, len(COLS) + 2,
             "Baseline = primary planned lanes through the SAME capacity-constrained solver. "
             "No primary lanes exist under disruption scenarios.").font = AR

    # ---------------- Summary (live formulas over the score sheets) ----------------
    wm = wb.create_sheet("Summary", 0)
    wm.cell(1, 1, "ChainLab - Official normalised 40/40/20 MinScore (cost/kg, delivery grain; "
                  "lower = better). Averages are over SCORED deliveries - unserved deliveries "
                  "are listed in Unassigned, not silently dropped.").font = H1
    scs = list(per_scenario)
    hdr = ["Metric"] + scs
    for j, h in enumerate(hdr, 1):
        c = wm.cell(3, j, h); c.font = BD; c.fill = GRY
    mcol = L["DeliveryMinScore"]
    rng = {sc: f"DeliveryScores!${mcol}${a}:${mcol}${b}" for sc, (a, b) in blocks_o.items()}
    stats = [("Average MinScore (scored deliveries)", "AVERAGE"), ("Median", "MEDIAN"),
             ("Std dev", "STDEV"), ("Best (min)", "MIN"), ("Worst (max)", "MAX")]
    r = 4
    avg_cell = {}
    for name, fn in stats:
        wm.cell(r, 1, name).font = AR
        for j, sc in enumerate(scs, 2):
            c = wm.cell(r, j, f"={fn}({rng[sc]})"); c.number_format = FMT; c.font = AR
            if fn == "AVERAGE":
                avg_cell[sc] = f"${get_column_letter(j)}$4"
        r += 1
    counts = [("Shipments routed / universe",
               lambda p: f"{p['routed']} / {p['universe']}"),
              ("Deliveries scored / total", lambda p: f"{p['scored']} / {p['total']}"),
              ("Multi-week routes selected", lambda p: p["multiweek"]),
              ("Unassigned shipments (see Unassigned)", lambda p: p["unassigned"])]
    for name, get in counts:
        wm.cell(r, 1, name).font = AR
        for j, sc in enumerate(scs, 2):
            wm.cell(r, j, get(per_scenario[sc])).font = AR
        r += 1

    r += 1
    wm.cell(r, 1, "BASELINE (primary planned lanes, same solver & normalisation)").font = BD; r += 1
    if "Normal" in blocks_b:
        a, b = blocks_b["Normal"]
        brng = f"BaselineScores!${mcol}${a}:${mcol}${b}"
        base_rows = [("Baseline average (Normal)", f"=AVERAGE({brng})"),
                     ("Baseline Q1 - excellent threshold", f"=QUARTILE({brng},1)"),
                     ("Baseline Q3 - weak warning threshold", f"=QUARTILE({brng},3)")]
        cells = {}
        for name, f in base_rows:
            wm.cell(r, 1, name).font = AR
            c = wm.cell(r, 2, f); c.number_format = FMT; c.font = AR
            cells[name.split(" -")[0].split(" (")[0]] = f"$B${r}"
            r += 1
        av = avg_cell["Normal"]
        stmts = [
            ("Beats baseline?",
             f'=IF({av}<{cells["Baseline average"]},"YES - average below baseline","NO")'),
            ("Excellent (at/below baseline Q1)?",
             f'=IF({av}<={cells["Baseline Q1"]},"YES - at/below Q1","NO - above Q1")'),
            ("Weak warning (above baseline Q3)?",
             f'=IF({av}>{cells["Baseline Q3"]},"FLAG - above Q3","OK - not above Q3")'),
            ("Same-population check (common deliveries)",
             f"opt {per_scenario['Normal']['opt_avg_common']} vs baseline "
             f"{per_scenario['Normal']['base_avg_common']} on "
             f"{per_scenario['Normal']['common_n']} deliveries served by both"),
        ]
        for name, f in stmts:
            wm.cell(r, 1, name).font = BD; wm.cell(r, 2, f).font = BD; r += 1
    wm.cell(r, 1, "Baseline under disruption: no primary planned lanes exist in "
                  "PrimaryHubDown / AirCapacityReduced - the optimiser finds alternatives "
                  "(see scenario columns); baseline is reported as unavailable, not zero.").font = AR
    r += 2
    for k, v in _run_config_rows(hub_disruption, scenarios or scs):
        wm.cell(r, 1, k).font = BD
        c = wm.cell(r, 2, v); c.font = AR; c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    wm.cell(r, 1, "Tradeoff explanation: see Assumptions sheet.").font = AR
    wm.column_dimensions["A"].width = 44
    for col in "BCD":
        wm.column_dimensions[col].width = 22

    # ---------------- RouteDecisions + Unassigned ----------------
    def df_sheet(name, df):
        ws = wb.create_sheet(name)
        for j, h in enumerate(df.columns, 1):
            c = ws.cell(1, j, h); c.font = BD; c.fill = GRY
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, v in enumerate(row, 1):
                ws.cell(i, j, None if pd.isna(v) else v).font = AR
        ws.freeze_panes = "A2"
    rk = ["scenario", "ShipmentID", "RouteOptionID", "FromHub", "ToHub", "TransportMode",
          "IsPrimary", "Qty", "WeeksRequired", "MultiWeek", "BaseLeadTimeDays",
          "EffectiveLeadTimeDays", "BaseCostEUR", "RiskScore", "CO2Kg"]
    full = pd.concat(route_frames, ignore_index=True)
    df_sheet("RouteDecisions", full[[c for c in rk if c in full.columns]])
    df_sheet("Unassigned", pd.DataFrame(unassigned_rows))

    del wb["Sheet"]
    wb.save(out_path)


def run_official_delivery(sheets, horizon_weeks, scenarios, out_path):
    dbs, E = deliveries_by_shipment(sheets)
    scaler = fit_scaler_delivery(sheets, horizon_weeks, dbs)
    qty = sheets["internal"].set_index("ShipmentID")["Qty"].to_dict()
    ACTIONS = {
        "handling_infeasible": "No handling-compatible hub - source alternate hub / supplier exception",
        "zero_capacity":       "Only routes pass a zero-capacity hub - free capacity or reroute",
        "capacity_escalation": "Exceeds quarter horizon - split order across periods or add capacity",
        "capacity_contention": "Weekly capacity contention - reschedule to another week"}
    summaries, per_scenario, route_frames, unassigned_rows = [], {}, [], []
    for sc in scenarios:
        res = solve_official_delivery(sc, sheets, scaler, horizon_weeks, dbs)
        base = solve_official_delivery(sc, sheets, scaler, horizon_weeks, dbs,
                                       restrict_primary=True)
        ds, bds = res["delivery_scores"], base["delivery_scores"]
        common = (set(zip(ds["DeliveryNo"], [sc] * len(ds)))
                  & set(zip(bds["DeliveryNo"], [sc] * len(bds)))) if len(bds) else set()
        common_dn = {d for d, _ in common}
        opt_c = ds[ds["DeliveryNo"].isin(common_dn)]["DeliveryMinScore"]
        base_c = bds[bds["DeliveryNo"].isin(common_dn)]["DeliveryMinScore"] if len(bds) else pd.Series(dtype=float)

        # unassigned detail (exported to the Unassigned sheet)
        annotated = add_capacity_fields(
            bc.apply_capability(bc.build_candidates(sheets, sc)), qty, horizon_weeks)
        annotated = annotated[annotated["ShipmentID"].isin(set(dbs))]
        real = annotated[annotated["BottleneckCapacityPerWeek"] > 0]
        best_real = (real.sort_values("EffectiveLeadTimeDays").groupby("ShipmentID").first()
                     if len(real) else pd.DataFrame())
        for sid in res["unassigned_ids"]:
            rk_ = res["reason_by_ship"].get(sid, "capacity_contention")
            unassigned_rows.append({
                "scenario": sc, "ShipmentID": sid, "Reason": rk_,
                "LinkedDeliveries": len(dbs.get(sid, [])),
                "BestAvailableWeeks": (int(best_real.loc[sid, "WeeksRequired"])
                                       if len(best_real) and sid in best_real.index else None),
                "BestAvailableRoute": (best_real.loc[sid, "RouteOptionID"]
                                       if len(best_real) and sid in best_real.index else None),
                "RecommendedAction": ACTIONS[rk_]})

        chosen = res["chosen"].copy(); chosen.insert(0, "scenario", sc); route_frames.append(chosen)
        per_scenario[sc] = {
            "ds": ds, "base_ds": bds,
            "routed": len(res["chosen"]), "universe": len(res["shipments"]),
            "scored": int(ds["DeliveryMinScore"].notna().sum()), "total": len(E),
            "multiweek": int(res["chosen"]["MultiWeek"].sum()) if len(res["chosen"]) else 0,
            "unassigned": len(res["unassigned_ids"]),
            "common_n": len(common_dn),
            "opt_avg_common": _r(opt_c.mean()) if len(common_dn) else None,
            "base_avg_common": _r(base_c.mean()) if len(common_dn) else None}
        summaries.append({
            "scenario": sc, "objective": "official cost/kg (delivery grain)",
            "solver": res["status"],
            "shipments_routed": len(res["chosen"]), "shipments_universe": len(res["shipments"]),
            "deliveries_scored": int(ds["DeliveryMinScore"].notna().sum()), "deliveries_total": len(E),
            "avg_delivery_MinScore": _r(ds["DeliveryMinScore"].mean()),
            "median": _r(ds["DeliveryMinScore"].median()), "std": _r(ds["DeliveryMinScore"].std()),
            "min": _r(ds["DeliveryMinScore"].min()), "max": _r(ds["DeliveryMinScore"].max()),
            "base_routed": len(base["chosen"]),
            "base_deliveries_scored": int(bds["DeliveryMinScore"].notna().sum()) if len(bds) else 0,
            "base_avg": _r(bds["DeliveryMinScore"].mean()) if len(bds) else None,
            "base_Q1": _r(bds["DeliveryMinScore"].quantile(.25)) if len(bds) else None,
            "base_Q3": _r(bds["DeliveryMinScore"].quantile(.75)) if len(bds) else None})
        print(f"\n=== {sc}  [OFFICIAL cost/kg, delivery grain] ===  (solver {res['status']})")
        print(f"  optimised: routed {len(res['chosen'])}/{len(res['shipments'])} shipments -> "
              f"scored {int(ds['DeliveryMinScore'].notna().sum())}/{len(E)} deliveries  "
              f"avg MinScore {_r(ds['DeliveryMinScore'].mean())}")
        if len(bds):
            print(f"  baseline : routed {len(base['chosen'])} -> scored "
                  f"{int(bds['DeliveryMinScore'].notna().sum())} deliveries  avg {_r(bds['DeliveryMinScore'].mean())}  "
                  f"Q1 {_r(bds['DeliveryMinScore'].quantile(.25))}  Q3 {_r(bds['DeliveryMinScore'].quantile(.75))}")
            print(f"  common {len(common_dn)} deliveries -> opt {per_scenario[sc]['opt_avg_common']} "
                  f"vs base {per_scenario[sc]['base_avg_common']}")
        else:
            print("  baseline : no primary planned lanes in this scenario (reported as unavailable)")

    _export_official(out_path, scaler, per_scenario, unassigned_rows, route_frames,
                     hub_disruption=sheets.get("hub_disruption"), scenarios=scenarios)
    print(f"\nWrote {out_path}")
    return summaries


def run_objective(sheets, horizon_weeks, scenarios, out_path):
    scaler = fit_scaler(sheets, horizon_weeks)
    label = "PROXY cost/piece score"
    summaries, route_frames, unassigned_rows = [], [], []
    for sc in scenarios:
        out = run_scenario(sc, sheets, scaler, horizon_weeks)
        s = out["summary"]; summaries.append(s); unassigned_rows.extend(out["unassigned_detail"])
        print(f"\n=== {sc}  [{label}] ===  (solver {s['solver']}, universe {s['universe']})")
        print(f"  optimised : solved {s['opt_solved']}/{s['universe']}  unassigned {s['opt_unassigned']}  "
              f"avg {s['opt_avg_MinScore']}  (min {s['opt_min']}, max {s['opt_max']})  mw {s['multiweek_selected']}")
        if s["base_solved"]:
            print(f"  baseline  : solved {s['base_solved']}/{s['universe']}  "
                  f"same {s['common_n']} -> opt {s['opt_avg_common']} vs base {s['base_avg_common']}")
        else:
            print(f"  baseline  : INFEASIBLE (no primary lanes in scenario)")
        print(f"  penalised objective -> opt {s['opt_penalised_obj']}  base {s['base_penalised_obj']}")
        ch = out["result"]["chosen"].copy(); ch.insert(0, "scenario", sc); route_frames.append(ch)

    with pd.ExcelWriter(out_path) as xw:
        pd.DataFrame(_run_config_rows(sheets.get("hub_disruption"), scenarios),
                     columns=["Setting", "Value"]).to_excel(
            xw, sheet_name="RunConfig", index=False)
        pd.DataFrame(summaries).to_excel(xw, sheet_name="Summary", index=False)
        keep = ["scenario", "ShipmentID", "MaterialFamily", "StageFrom", "StageTo",
                "RouteOptionID", "FromHub", "ToHub", "TransportMode", "IsPrimary", "Qty",
                "ShipmentWeight", "CostForScore", "BottleneckCapacityPerWeek", "WeeksRequired",
                "MultiWeek", "BaseLeadTimeDays", "EffectiveLeadTimeDays", "BaseCostEUR",
                "RiskScore", "CO2Kg", "WeightedScore"]
        full = pd.concat(route_frames, ignore_index=True)
        full[[c for c in keep if c in full.columns]].to_excel(xw, sheet_name="SelectedRoutes", index=False)
        pd.DataFrame(unassigned_rows).to_excel(xw, sheet_name="Unassigned", index=False)
    print(f"\nWrote {out_path}")
    return summaries


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--scenario", default="all")
    ap.add_argument("--objective", default="both", choices=["official", "proxy", "both"],
                    help="official = cost/kg (132 linked); proxy = cost/piece (all 240)")
    ap.add_argument("--horizon-weeks", type=int, default=PLANNING_HORIZON_WEEKS)
    ap.add_argument("--hub-disruption", default=None,
                    help="Hub-side capacity-cut axis (guide Scenario 1): "
                         "omit for clean network; 'Port congestion', "
                         "'Labor shortage', 'Weather disruption', or "
                         "'all' (legacy pre-18-Jul behaviour).")
    args = ap.parse_args()

    sheets = bc.load_sheets(hub_disruption=args.hub_disruption)
    scen = SCENARIOS if args.scenario == "all" else [args.scenario]

    tag = ("" if args.hub_disruption is None
           else "_" + args.hub_disruption.lower().replace(" ", ""))
    print(f"Hub disruption axis: {args.hub_disruption or 'None (clean network)'}")
    if args.objective in ("official", "both"):
        run_official_delivery(sheets, args.horizon_weeks, scen,
                              f"optimizer_combined_extension{tag}.xlsx")
    if args.objective in ("proxy", "both"):
        run_objective(sheets, args.horizon_weeks, scen,
                      f"optimizer_proxy_resilience{tag}.xlsx")
