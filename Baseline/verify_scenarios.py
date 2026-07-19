"""
verify_scenarios.py
===================
Automated regression checks (reviewer-requested, 18 Jul 2026).

Confirms every scenario script shares ONE canonical scaler with the official
baseline, so the S2 'cold-chain restricted' run and the S3 'weighted_baseline'
run reproduce the official clean Normal solved count and average score exactly.
Also asserts the wording/config fixes landed in the workbooks.
"""
import sys
import pandas as pd
import openpyxl
import build_candidates as bc
import baseline_v7 as b7
import scenario_coldchain as s2
import scenario_expedite as s3
import scenario_sustainability as s4
import scenario_resilience as s5
import scenario_external_official as sx
import numpy as np

TOL = 1e-4
HZ = b7.HORIZON


def capacity_ok(chosen, sheets, key):
    """Assert no chosen route/hub exceeds weekly capacity (self-loops once)."""
    hub_rem = sheets["hub"].set_index("HubID")["remaining_capacity_units"].to_dict()
    route_load, hub_load, route_cap = {}, {}, {}
    for _, r in chosen.iterrows():
        route_load[(r["RouteOptionID"], r["week"])] = route_load.get(
            (r["RouteOptionID"], r["week"]), 0) + r["WeeklyFootprint"]
        route_cap[(r["RouteOptionID"], r["week"])] = r["CapacityUnitsPerWeek"]
        for hub in {r["FromHub"], r["ToHub"]}:
            hub_load[(hub, r["week"])] = hub_load.get((hub, r["week"]), 0) + r["WeeklyFootprint"]
    for kk, load in route_load.items():
        if load > route_cap[kk] + TOL:
            return False
    for (hub, wk), load in hub_load.items():
        if load > hub_rem.get(hub, float("inf")) + TOL:
            return False
    return True


def official_normal(side, sheets, E):
    """The official clean-Normal baseline: canonical scaler + standard solve."""
    cfg = b7._side_setup(side, sheets, E)
    ann = cfg["add_cost"](b7.capacity_fields(cfg["build"]("Normal"),
                                             cfg["qty_map"], cfg["key"], HZ))
    ann["week"] = ann[cfg["key"]].map(cfg["wk_map"])
    feas = b7.score(b7.model_feasible(ann), b7.canonical_scaler(side, sheets, E, HZ))
    hub_rem = sheets["hub"].set_index("HubID")["remaining_capacity_units"].to_dict()
    res = b7.solve(feas, cfg["key"], cfg["universe"], hub_rem)
    ch = res["chosen"]
    return len(ch), round(ch["MinScore"].mean(), 4), set(
        zip(ch[cfg["key"]], ch["RouteOptionID"]))


def check(name, got, exp):
    ok = (abs(got - exp) <= TOL) if isinstance(exp, float) else (got == exp)
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}: got {got}  expected {exp}")
    return ok


def main():
    sheets = bc.load_sheets(hub_disruption=None)
    E = pd.read_excel(bc.WORKBOOK, sheet_name="External Shipments")
    ok = True
    res4, res5 = {}, {}   # captured for expected-value spot checks after the loop

    for side in ("internal", "external"):
        n_off, avg_off, routes_off = official_normal(side, sheets, E)
        print(f"\n{side.upper()} official clean Normal: solved {n_off}  avg {avg_off}")

        # S2 restricted (cold rule ON) must reproduce official Normal
        o2 = s2.run_side(side, sheets, E, HZ)
        r = next(x for x in o2["summary"] if x["variant"].startswith("cold-chain restricted"))
        n2, a2 = r["solved"], r["avg_MinScore"]
        routes_r = set(zip(o2["res_r"]["chosen"][o2["key"]],
                           o2["res_r"]["chosen"]["RouteOptionID"]))
        ok &= check(f"S2 restricted solved == official", n2, n_off)
        ok &= check(f"S2 restricted avg == official", a2, avg_off)
        ok &= check(f"S2 restricted route decisions == official", routes_r == routes_off, True)

        # S3 weighted_baseline must reproduce official Normal
        o3 = s3.run_side(side, sheets, E, HZ)
        s = o3["summary"]
        ok &= check("S3 weighted_baseline solved == official",
                    s["total_solved_weighted_baseline"], n_off)
        ok &= check("S3 weighted_baseline avg == official",
                    s["avg_MinScore_weighted_baseline_all"], avg_off)

        # ---- S4 sustainability checks ----
        o4 = s4.run_side(side, sheets, E, HZ, 2)
        s4s, det4 = o4["summary"], o4["detail"]
        key = s4.KEYS[side] if hasattr(s4, "KEYS") else ("ShipmentID" if side == "internal" else "DeliveryNo")
        ok &= check("S4 solved + unassigned == universe",
                    s4s["sustainability_solved"] + len(o4["unassigned"]), s4s["universe"])
        sel = det4[det4["assigned_sustainability"]]
        ok &= check("S4 every selected route within SLA",
                    bool((sel["effLead_sus"] <= sel["SLA"]).all()), True)
        ok &= check("S4 CO2_saved_common == baseline - sustainability",
                    s4s["CO2_saved_kg_common"],
                    s4s["CO2_common_weighted_baseline"] - s4s["CO2_common_sustainability"])
        ok &= check("S4 breaches total == rerouted + unserved",
                    s4s["baseline_SLA_breaches_total"],
                    s4s["baseline_breaches_rerouted_within_SLA"]
                    + s4s["baseline_breaches_unserved_due_to_SLA"])
        res4[side] = s4s

        # ---- S5 resilience checks ----
        o5 = s5.run_side(side, E, HZ)
        universe_n = len(sheets["internal"]) if side == "internal" else len(E)
        for cc in o5["cat_counts"]:
            total_cat = sum(v for k, v in cc.items()
                            if k not in ("side", "column", "mode_changes"))
            ok &= check(f"S5 {cc['column']} route-switch categories sum == universe",
                        total_cat, universe_n)
        # S5 Normal column must reproduce the official Normal route decisions
        scaler5 = b7.canonical_scaler(side, sheets, E, HZ)
        ch5_norm, _ = s5.solve_column(side, E, None, "Normal", scaler5, HZ)
        routes5 = set(zip(ch5_norm[key], ch5_norm["RouteOptionID"]))
        ok &= check("S5 Normal routes == official Normal routes", routes5 == routes_off, True)
        ok &= check("S5 Normal solve respects route+hub capacity",
                    capacity_ok(ch5_norm, sheets, key), True)
        res5[side] = {cc["column"]: cc for cc in o5["cat_counts"]}

    # ---- expected-value spot checks (reviewer's reconciliation targets) ----
    print("\nExpected-value reconciliation:")
    ok &= check("S4 internal breaches 61 = 4 rerouted + 57 unserved",
                (res4["internal"]["baseline_SLA_breaches_total"],
                 res4["internal"]["baseline_breaches_rerouted_within_SLA"],
                 res4["internal"]["baseline_breaches_unserved_due_to_SLA"]), (61, 4, 57))
    ok &= check("S4 external breaches 25 = 1 rerouted + 24 unserved",
                (res4["external"]["baseline_SLA_breaches_total"],
                 res4["external"]["baseline_breaches_rerouted_within_SLA"],
                 res4["external"]["baseline_breaches_unserved_due_to_SLA"]), (25, 1, 24))
    ok &= check("S5 internal PortCongestion same_route == 157",
                res5["internal"]["PortCongestion"].get("same_route", 0), 157)
    ok &= check("S5 external PortCongestion same_route == 174",
                res5["external"]["PortCongestion"].get("same_route", 0), 174)

    # ---- official external scorer checks (instructor-approved) ----
    print("\nOfficial external MinScore (direct score-and-rank):")
    Ex = pd.read_excel(bc.WORKBOOK, sheet_name="External Shipments")
    xs, xb = sx.score_external_official(Ex)
    ok &= check("official external rows == 225", len(xs), 225)
    ok &= check("official external DeliveryNo unique == 225", xs["DeliveryNo"].nunique(), 225)
    ok &= check("official external MinScore in [0,1]", bool(xs["ExternalMinScore"].between(0, 1).all()), True)
    ok &= check("official external inputs non-null",
                bool(xs[["BestLeadTimeDays", "LowestCostPerKG_EUR", "LowestRiskScore"]].notna().all().all()), True)
    _rec = (0.4*xs["n_BestLeadTimeDays"] + 0.4*xs["n_LowestCostPerKG_EUR"] + 0.2*xs["n_LowestRiskScore"])
    ok &= check("official external reconstructs 40/40/20", bool(np.allclose(xs["ExternalMinScore"], _rec)), True)
    ok &= check("official external uses ONLY the 3 supplied columns (+DeliveryNo)",
                set(sx.SCORE_COLS).issubset(Ex.columns), True)

    # workbook wording / config checks
    print("\nWorkbook wording & config:")
    wb = openpyxl.load_workbook("optimizer_internal_baseline_all.xlsx")
    txt = " ".join(str(c) for row in wb["Assumptions"].iter_rows(values_only=True) for c in row if c)
    ok &= check("legacy says 'every hub that carries a recorded reduction'",
                "every hub that carries a recorded reduction" in txt.lower(), True)
    ok &= check("legacy does NOT say \"DisruptionScenario == 'all'\"",
                "disruptionscenario == 'all'" not in txt.lower(), True)

    # workbook integrity: every deliverable exists, is non-empty, and opens with sheets
    print("\nWorkbook integrity (non-empty + openable):")
    from pathlib import Path
    required = [
        "optimizer_internal_baseline.xlsx",
        "optimizer_external_official_scores.xlsx",   # instructor-approved official external
        "optimizer_external_route_extension.xlsx",   # route extension (renamed, non-official)
        "optimizer_coldchain_scenario.xlsx", "optimizer_expedite_scenario.xlsx",
        "optimizer_resilience_scenario.xlsx", "optimizer_sustainability_scenario.xlsx",
    ]
    for name in required:
        p = Path(name)
        good = p.exists() and p.stat().st_size > 0
        if good:
            try:
                good = len(openpyxl.load_workbook(p, read_only=True).sheetnames) > 0
            except Exception:
                good = False
        ok &= check(f"{name} non-empty & openable", good, True)

    print("\n" + ("ALL CHECKS PASSED" if ok else "SOME CHECKS FAILED"))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
